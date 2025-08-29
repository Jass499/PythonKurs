"""
    Autor:      Robert Panzirsch
    Datum:      x.x.20xx
    Version:    1.0

    Excelmappe im xlsx Format mit Paket openpyxl schreiben
    Achtung: Bei Formeln und Zahlenformat etc: Culture: en-us

"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from random import seed, randint  # Zufallsmodul um Testdaten zu erstellen

arbeitsmappe = "datentransfer.xlsx"
arbeitsblatt1 = "Blatt1"
arbeitsblatt2 = "Blatt2"
zelle1 = "C3"
zelle2 = "C4"
position1 = "G2"
position2 = "G22"
wert = "Datentransfer nach Excel ..."
formel = "=today()"

mappe = Workbook()
# Objekt der Workbook Klasse instanzieren, enthält schon ein Blatt
blatt1 = mappe.active
# Zugriff auf aktives, erstes Blatt schaffen
blatt1.title = arbeitsblatt1
# Blatt beschriften

blatt1[zelle1] = wert
# Wert schreiben
blatt1[zelle2] = formel
# Formel schreiben
blatt1[zelle2].number_format = "DD/MM/YYYY"
# Zahlenformat einstellen
blatt1.column_dimensions[zelle1[0]].width = 25
# Spaltenbreite in Spalte C einstellen

blatt2 = mappe.create_sheet(arbeitsblatt2)
# füge ein zweites Blatt in die Mappe und beschrifte

# Feldnamen und Schlüssel eintragen
for spaltennr in range(1, 6):
    blatt2.cell(row=1, column=spaltennr, value="Feld" + chr(64 + spaltennr))
for zeilennr in range(2, 21):
    blatt2.cell(row=zeilennr, column=1, value="Schluessel" + str(zeilennr - 1))

# zufällige Umsätze mit Buchhaltungszahlenformat eintragen
seed()
for zeilennr in range(2, 21):
    # Zeile 2 bis 20
    for spaltennr in range(2, 6):
        # Spalte 2 bis 5
        zelle = blatt2.cell(row=zeilennr, column=spaltennr, value=randint(1, 100) * 100)
        zelle.number_format = "_-\"€\" * #,##0.00_-"

# Spaltenbreiten einstellen
for spalte in blatt2.columns:
    breite = 0
    spaltenbuchstabe = spalte[0].column_letter
    # Spaltenbuchstaben erhalten
    for zelle in spalte:
        if len(str(zelle.value)) > breite:
            breite = len(zelle.value) + 8
    blatt2.column_dimensions[spaltenbuchstabe].width = breite

# Erstellen eines Säulendiagramms
diagramm1 = BarChart()
# Diagramm Objekt instanzieren
diagramm1.title = "Umsatzdiagramm"
diagramm1.x_axis.title = "Schlüssel"
diagramm1.y_axis.title = "Umsatz"
diagramm1.x_axis.delete = False
diagramm1.y_axis.delete = False
# Achsen sichtbar
diagramm1.y_axis.scaling.min = 0
daten1 = Reference(blatt2, min_col=2, min_row=1, max_col=5, max_row=20)
kategorien1 = Reference(blatt2, min_col=1, min_row=2, max_row=20)
diagramm1.add_data(daten1, titles_from_data=True)
diagramm1.set_categories(kategorien1)
diagramm1.width = 20
diagramm1.height = 10
diagramm1.legend.position = "r"
diagramm1.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.1, h=0.8, w=0.8))
blatt2.add_chart(diagramm1, position1)

# Erstellen eines Liniendiagramms
diagramm2 = LineChart()
# Diagramm Objekt instanzieren
diagramm2.title = "Umsatzverlauf"
diagramm2.x_axis.title = "Schlüsselinformation"
diagramm2.y_axis.title = "Umsatzinformation"
diagramm2.x_axis.delete = False
diagramm2.y_axis.delete = False
# Achsen sichtbar
diagramm2.y_axis.scaling.min = 0
daten2 = Reference(blatt2, min_col=2, min_row=1, max_col=5, max_row=20)
kategorien2 = Reference(blatt2, min_col=1, min_row=2, max_row=20)
diagramm2.add_data(daten2, titles_from_data=True)
diagramm2.set_categories(kategorien2)
diagramm2.width = 20
diagramm2.height = 10
diagramm2.legend.position = "r"
diagramm2.layout = Layout(manualLayout=ManualLayout(yMode="edge", xMode="edge", x=0.1, y=0.1, h=0.8, w=0.8))
blatt2.add_chart(diagramm2, position2)

mappe.save(filename=arbeitsmappe)
# speichern
mappe.close()

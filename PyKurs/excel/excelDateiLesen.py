"""
    Autor:      Jasmin Schmitt
    Datum:      29.08.2025
    Version:    1.0

    Excelmappe im xlsx Format mit Modul openpyxl auslesen

"""

from openpyxl import load_workbook

arbeitsmappe = "datenbank.xlsx"
arbeitsblatt = "Tabelle1"
zelle = "C3"

mappe = load_workbook(arbeitsmappe)
blatt = mappe[arbeitsblatt]

for zeilennr in range(1, 6):
    for spaltennr in range(1, 7):
        print((str(blatt.cell(row=zeilennr, column=spaltennr).value) + "           ")[:11], end="")
    print()

print("Die Zelle", zelle, "enthält den Wert:", blatt[zelle].value)
mappe.close()
